# -*- coding: utf-8 -*-
"""엑셀 '정부지원사업 모니터링 사이트목록 통합본'의 사이트를 sites.json에 중복 제외 병합.

사용:
  python scripts/merge_sites_from_xlsx.py            # 드라이런(요약만 출력)
  python scripts/merge_sites_from_xlsx.py --write     # sites.json 실제 갱신
"""
from __future__ import annotations
import os, sys, json, hashlib, re, shutil, tempfile
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SITES_PATH = ROOT / "sites.json"
XLSX = Path(r"C:\Users\ekth3\OneDrive\문서\카카오톡 받은 파일\정부지원사업_모니터링_사이트목록_통합본_342개 (1).xlsx")
SHEET = "전체목록"


def open_xlsx():
    """OneDrive/Excel 잠금 회피: 임시 복사 후 읽기."""
    try:
        return openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / "_sites_src_merge.xlsx"
        shutil.copy2(XLSX, tmp)
        return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def norm_url(u: str) -> str:
    """중복 판정용 URL 정규화: scheme/host 소문자, 마지막 슬래시 제거, 쿼리 유지."""
    if not u:
        return ""
    u = u.strip()
    try:
        p = urlsplit(u)
    except ValueError:
        return u.lower()
    scheme = (p.scheme or "http").lower()
    netloc = p.netloc.lower()
    path = p.path.rstrip("/")
    # www. 제거(동일 사이트 변형 흡수)
    if netloc.startswith("www."):
        netloc = netloc[4:]
    return urlunsplit((scheme, netloc, path, p.query, ""))


def make_id(url: str, taken: set[str]) -> str:
    base = "imp_" + hashlib.md5(url.encode("utf-8")).hexdigest()[:8]
    cid = base
    i = 1
    while cid in taken:
        cid = f"{base}_{i}"
        i += 1
    taken.add(cid)
    return cid


def main() -> int:
    write = "--write" in sys.argv

    sites = json.loads(SITES_PATH.read_text(encoding="utf-8"))
    existing_norm = {norm_url(s.get("url", "")) for s in sites if s.get("url")}
    existing_ids = {s.get("id", "") for s in sites}

    wb = open_xlsx()
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # 컬럼: 번호, 사이트명, 게시판명, 모집빈도, URL, 우선순위, 모니터링 주기, 비고

    added: list[dict] = []
    seen_new: set[str] = set()
    dup_existing = 0
    dup_within = 0
    skipped_nourl = 0

    for r in rows:
        name = (r[1] or "").strip() if r[1] else ""
        board = (r[2] or "").strip() if r[2] else ""
        url = (r[4] or "").strip() if r[4] else ""
        prio = (r[5] or "").strip() if r[5] else ""
        cycle = (r[6] or "").strip() if r[6] else ""
        memo = (r[7] or "").strip() if r[7] else ""

        if not url or not re.match(r"^https?://", url, re.I):
            skipped_nourl += 1
            continue
        nu = norm_url(url)
        if nu in existing_norm:
            dup_existing += 1
            continue
        if nu in seen_new:
            dup_within += 1
            continue
        seen_new.add(nu)

        note_parts = [p for p in (board, cycle and f"주기:{cycle}", prio and f"우선순위:{prio}", memo) if p]
        entry = {
            "id": make_id(url, existing_ids),
            "name": name or url,
            "type": "html_table",
            "url": url,
            "enabled": True,
            "is_aggregator": False,
            "note": " / ".join(note_parts),
        }
        added.append(entry)

    print(f"기존 sites.json: {len(sites)}개")
    print(f"엑셀 '{SHEET}' 데이터행: {len(rows)}개")
    print(f"  - 기존과 중복(제외): {dup_existing}")
    print(f"  - 엑셀 내부 중복(제외): {dup_within}")
    print(f"  - URL 없음/형식오류(제외): {skipped_nourl}")
    print(f"  - 신규 추가: {len(added)}")
    print(f"병합 후 총합: {len(sites) + len(added)}개")

    if write:
        merged = sites + added
        SITES_PATH.write_text(
            json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        print(f"\n[WRITE] sites.json 갱신 완료 → {len(merged)}개")
    else:
        print("\n[DRY-RUN] --write 미지정. 파일 변경 없음.")
        for e in added[:5]:
            print("  예시:", json.dumps(e, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
